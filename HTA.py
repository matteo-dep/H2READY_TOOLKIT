import streamlit as st
import pandas as pd
from io import BytesIO
import requests
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURAZIONE PAGINA E LINGUA
# ==========================================
st.set_page_config(page_title="H2READY TOOLKIT - Tool 2.1", layout="wide")

LANG_OPTIONS = {"Italiano": "it", "English": "en", "Slovenščina": "sl"}
lang_choice = st.sidebar.selectbox("🌐 Lingua / Language / Jezik", list(LANG_OPTIONS.keys()))
LANG = LANG_OPTIONS[lang_choice]

# ==========================================
# 2. SISTEMA DI CLASSIFICAZIONE PER LINGUA
#    IT -> ATECO | EN -> NACE | SL -> SKD
#    (Tutti e tre coincidono fino alla 4a cifra: NACE Rev.2)
# ==========================================
# Nome del sistema di classificazione mostrato all'utente in base alla lingua
CLASSIF_NAME = {"it": "ATECO", "en": "NACE", "sl": "SKD"}

# --- DIZIONARIO CODICI HTA (4 cifre - Hard to Abate e RED III) ---
# La CHIAVE (4 cifre) è identica nei 3 sistemi. Cambia solo la lingua della descrizione.
CODE_DESCRIPTIONS = {
    "it": {
        '1910': 'Prodotti della cokeria (1000-1100°C)',
        '1920': 'Raffinazione di prodotti petroliferi (500°C)',
        '2011': 'Produzione di gas industriali - SMR (700°C)',
        '2012': 'Produzione di coloranti e pigmenti (600-1000°C)',
        '2013': 'Chimica inorganica di base (500°C)',
        '2014': 'Chimica organica di base - Cracking (700-1100°C)',
        '2015': 'Fabbricazione di fertilizzanti e composti azotati (700°C)',
        '2016': 'Materie plastiche primarie (700-1100°C)',
        '2311': 'Fabbricazione di vetro piano (~1500°C)',
        '2313': 'Fabbricazione di vetro cavo (~1500°C)',
        '2320': 'Fabbricazione di prodotti refrattari (1200-1600°C)',
        '2332': 'Fabbricazione di mattoni e tegole (900-1200°C)',
        '2351': 'Produzione di cemento (1400-1500°C)',
        '2352': 'Produzione di calce e gesso (900-1200°C)',
        '2410': 'Produzione di ferro, acciaio e ferroleghe (1200-1600°C)',
        '2431': 'Trafilatura a freddo (600-1000°C)',
        '2442': 'Produzione di alluminio e semilavorati (660°C)',
        '2443': 'Produzione di rame e semilavorati (1000-1200°C)',
        '2444': 'Produzione di altri metalli non ferrosi (400-1200°C)',
        '2451': 'Fusione di ghisa (1200-1400°C)',
        '2452': 'Fusione di acciaio (1400-1600°C)',
        '2453': 'Fusione di metalli leggeri (650-1650°C)',
        '2550': 'Fucinatura, stampaggio e profilatura metalli (900-1200°C)',
        '2561': 'Trattamento e rivestimento dei metalli (500-1100°C)',
        '2562': 'Lavorazioni meccaniche termiche (500°C)',
        '3511': 'Produzione energia elettrica da fonti fossili (600-1200°C)',
        '3530': 'Produzione di vapore industriale (500°C)',
        '3821': 'Trattamento rifiuti non pericolosi - Incenerimento (850-1100°C)',
        '3822': 'Trattamento rifiuti pericolosi - Incenerimento (1000-1200°C)',
        '3832': 'Recupero rottami metallici (600-1500°C)',
    },
    "en": {
        '1910': 'Coke oven products (1000-1100°C)',
        '1920': 'Refining of petroleum products (500°C)',
        '2011': 'Industrial gases - SMR (700°C)',
        '2012': 'Dyes and pigments (600-1000°C)',
        '2013': 'Basic inorganic chemicals (500°C)',
        '2014': 'Basic organic chemicals - Cracking (700-1100°C)',
        '2015': 'Fertilizers and nitrogen compounds (700°C)',
        '2016': 'Primary plastics (700-1100°C)',
        '2311': 'Flat glass manufacturing (~1500°C)',
        '2313': 'Hollow glass manufacturing (~1500°C)',
        '2320': 'Refractory products (1200-1600°C)',
        '2332': 'Bricks and tiles (900-1200°C)',
        '2351': 'Cement production (1400-1500°C)',
        '2352': 'Lime and plaster (900-1200°C)',
        '2410': 'Iron, steel and ferro-alloys (1200-1600°C)',
        '2431': 'Cold drawing of metal (600-1000°C)',
        '2442': 'Aluminium production (660°C)',
        '2443': 'Copper production (1000-1200°C)',
        '2444': 'Other non-ferrous metals (400-1200°C)',
        '2451': 'Casting of iron (1200-1400°C)',
        '2452': 'Casting of steel (1400-1600°C)',
        '2453': 'Casting of light metals (650-1650°C)',
        '2550': 'Forging, pressing, stamping of metal (900-1200°C)',
        '2561': 'Treatment and coating of metals (500-1100°C)',
        '2562': 'Machining / thermal works (500°C)',
        '3511': 'Electricity from fossil sources (600-1200°C)',
        '3530': 'Industrial steam production (500°C)',
        '3821': 'Non-hazardous waste incineration (850-1100°C)',
        '3822': 'Hazardous waste incineration (1000-1200°C)',
        '3832': 'Recovery of metal scrap (600-1500°C)',
    },
    "sl": {
        '1910': 'Proizvodnja koksa (1000-1100°C)',
        '1920': 'Predelava naftnih derivatov (500°C)',
        '2011': 'Proizvodnja industrijskih plinov - SMR (700°C)',
        '2012': 'Proizvodnja barvil in pigmentov (600-1000°C)',
        '2013': 'Proizvodnja anorganskih kemikalij (500°C)',
        '2014': 'Proizvodnja organskih kemikalij - kreking (700-1100°C)',
        '2015': 'Proizvodnja gnojil in dušikovih spojin (700°C)',
        '2016': 'Proizvodnja primarnih plastičnih mas (700-1100°C)',
        '2311': 'Proizvodnja ravnega stekla (~1500°C)',
        '2313': 'Proizvodnja votlega stekla (~1500°C)',
        '2320': 'Proizvodnja ognjevzdržnih izdelkov (1200-1600°C)',
        '2332': 'Proizvodnja opeke in strešnikov (900-1200°C)',
        '2351': 'Proizvodnja cementa (1400-1500°C)',
        '2352': 'Proizvodnja apna in mavca (900-1200°C)',
        '2410': 'Proizvodnja železa, jekla in ferozlitin (1200-1600°C)',
        '2431': 'Hladno vlečenje kovin (600-1000°C)',
        '2442': 'Proizvodnja aluminija (660°C)',
        '2443': 'Proizvodnja bakra (1000-1200°C)',
        '2444': 'Proizvodnja drugih neželeznih kovin (400-1200°C)',
        '2451': 'Litje železa (1200-1400°C)',
        '2452': 'Litje jekla (1400-1600°C)',
        '2453': 'Litje lahkih kovin (650-1650°C)',
        '2550': 'Kovanje, stiskanje in profiliranje kovin (900-1200°C)',
        '2561': 'Obdelava in prevlekanje kovin (500-1100°C)',
        '2562': 'Mehanska in toplotna obdelava (500°C)',
        '3511': 'Proizvodnja električne energije iz fosilnih virov (600-1200°C)',
        '3530': 'Oskrba s paro (500°C)',
        '3821': 'Sežig nenevarnih odpadkov (850-1100°C)',
        '3822': 'Sežig nevarnih odpadkov (1000-1200°C)',
        '3832': 'Predelava kovinskih odpadkov (600-1500°C)',
    },
}

# --- DIZIONARIO MACRO (2 cifre - Fallback "a cascata") ---
MACRO_DESCRIPTIONS = {
    "it": {
        '10': 'Industrie alimentari', '11': 'Industria delle bevande', '13': 'Industrie tessili',
        '14': 'Abbigliamento', '15': 'Articoli in pelle e cuoio', '16': 'Industria del legno',
        '17': 'Fabbricazione di carta', '18': 'Stampa e riproduzione', '19': 'Cokeria e Raffinazione',
        '20': 'Fabbricazione di prodotti chimici', '21': 'Prodotti farmaceutici', '22': 'Gomma e materie plastiche',
        '23': 'Minerali non metalliferi (Vetro, Cemento, Ceramica)', '24': 'Metallurgia di base',
        '25': 'Prodotti in metallo', '26': 'Computer e apparecchiature elettroniche',
        '27': 'Apparecchiature elettriche', '28': 'Macchinari e apparecchiature',
        '29': 'Autoveicoli e rimorchi', '30': 'Altri mezzi di trasporto',
        '31': 'Fabbricazione di mobili', '32': 'Altre industrie manifatturiere',
        '33': 'Riparazione e installazione macchine', '35': 'Energia elettrica, gas, vapore',
        '36': 'Raccolta e depurazione acque', '37': 'Gestione reti fognarie',
        '38': 'Trattamento e smaltimento rifiuti', '39': 'Risanamento e gestione rifiuti',
        '41': 'Costruzione di edifici', '42': 'Ingegneria civile', '43': 'Lavori di costruzione specializzati',
        '63': 'Servizi di informazione e Data Center',
    },
    "en": {
        '10': 'Food industries', '11': 'Beverage industry', '13': 'Textile industries',
        '14': 'Wearing apparel', '15': 'Leather products', '16': 'Wood industry',
        '17': 'Paper manufacturing', '18': 'Printing and reproduction', '19': 'Coke and refining',
        '20': 'Manufacture of chemicals', '21': 'Pharmaceutical products', '22': 'Rubber and plastics',
        '23': 'Non-metallic mineral products (Glass, Cement, Ceramics)', '24': 'Basic metals',
        '25': 'Fabricated metal products', '26': 'Computers and electronics',
        '27': 'Electrical equipment', '28': 'Machinery and equipment',
        '29': 'Motor vehicles and trailers', '30': 'Other transport equipment',
        '31': 'Furniture manufacturing', '32': 'Other manufacturing',
        '33': 'Repair and installation of machinery', '35': 'Electricity, gas, steam',
        '36': 'Water collection and treatment', '37': 'Sewerage',
        '38': 'Waste treatment and disposal', '39': 'Remediation activities',
        '41': 'Construction of buildings', '42': 'Civil engineering', '43': 'Specialised construction',
        '63': 'Information services and Data Centers',
    },
    "sl": {
        '10': 'Proizvodnja živil', '11': 'Proizvodnja pijač', '13': 'Proizvodnja tekstilij',
        '14': 'Proizvodnja oblačil', '15': 'Proizvodnja usnja in usnjenih izdelkov', '16': 'Obdelava in predelava lesa',
        '17': 'Proizvodnja papirja', '18': 'Tiskarstvo', '19': 'Proizvodnja koksa in naftnih derivatov',
        '20': 'Proizvodnja kemikalij', '21': 'Proizvodnja farmacevtskih izdelkov', '22': 'Izdelki iz gume in plastike',
        '23': 'Nekovinski mineralni izdelki (steklo, cement, keramika)', '24': 'Proizvodnja kovin',
        '25': 'Proizvodnja kovinskih izdelkov', '26': 'Računalniki in elektronske naprave',
        '27': 'Proizvodnja električnih naprav', '28': 'Proizvodnja strojev in naprav',
        '29': 'Proizvodnja motornih vozil in prikolic', '30': 'Proizvodnja drugih vozil in plovil',
        '31': 'Proizvodnja pohištva', '32': 'Druga predelovalna dejavnost',
        '33': 'Popravila in montaža strojev in naprav', '35': 'Oskrba z elektriko, plinom in paro',
        '36': 'Zbiranje, čiščenje in distribucija vode', '37': 'Ravnanje z odplakami',
        '38': 'Ravnanje z odpadki', '39': 'Saniranje okolja',
        '41': 'Gradnja stavb', '42': 'Gradnja inženirskih objektov', '43': 'Specializirana gradbena dela',
        '63': 'Informacijske storitve in podatkovni centri',
    },
}

# ==========================================
# 3. DIZIONARIO INTERFACCIA (i18n)
# ==========================================
T = {
    "it": {
        "title": "🚀 H2READY TOOLKIT - Tool 2.1: Scouting aziende HTA e RED III",
        "credits": "Sviluppato all'interno del progetto [INTERREG H2Ready](https://www.ita-slo.eu/en/h2ready) da **Matteo De Piccoli - [APE FVG](https://www.ape.fvg.it/)**",
        "instr_title": "📖 GUIDA OPERATIVA (Leggi prima di iniziare)",
        "instructions_md": """
### 🎯 Qual è il tuo obiettivo?
Come referente dell'Ente, il tuo compito è **mappare le industrie del tuo territorio** per capire quali hanno *reale* necessità di passare all'idrogeno verde (es. acciaierie, vetrerie, chimica).

**Segui questi 3 passaggi:**

**🟢 FASE 1: Screening Iniziale**
Carica il file di screening nella sezione dedicata. Il simulatore filtrerà le aziende idonee. Se non hai il file, scarica il template subito sotto l'area di caricamento.

**🟡 FASE 2: Consolidamento Fabbisogni**
Inserisci le stime dei fabbisogni (tonnellate/anno) delle aziende risultate idonee. Anche qui, trovi il template dedicato sotto l'uploader.

**🔵 FASE 3: Esportazione**
Inserisci il tuo **Codice Identificativo** in fondo alla pagina e clicca "Salva" per inviare i dati al database centrale.
        """,
        "note_codes": "**Nota sui codici ATECO:** il tool legge solo le **prime 4 cifre** del codice (la *Classe*), ignorando le ultime cifre che hanno finalità puramente contabili/statistiche. Se inserisci un codice non presente nel database prioritario per l'idrogeno, il tool riconosce comunque il **macro-settore** (prime 2 cifre) ed emette un alert chiedendoti di verificare il dato.",
        "info_template1": "💡 Se non hai ancora i dati di screening, usa questo modello. **Attenzione: i dati vanno caricati mantenendo esattamente questo formato (stessi nomi delle colonne e struttura).**",
        "info_template2": "💡 Usa questo modello per consolidare le tonnellate di idrogeno necessarie. **Attenzione: i dati vanno caricati rispettando rigorosamente il formato di questo template.**",
        "btn_template1": "📥 Scarica Template Screening (Fase 1)",
        "btn_template2": "📥 Scarica Template Fabbisogni (Fase 2)",
        "header_fase1": "📤 FASE 1: Caricamento e Analisi Termodinamica",
        "header_fase2": "📤 FASE 2: Consolidamento Fabbisogni",
        "input_id": "Inserisci il Codice Identificativo (es. 030043):",
        "btn_export": "🚀 Salva Risultati nel Database Centrale",
        "export_success": "✅ Dati salvati con successo nel Master Database!",
        "col_desc": "Descrizione settore",
        "col_score": "Punteggio H2",
        "col_verdict": "Esito termodinamico",
        "alert_codes": "⚠️ Codici fuori dal database prioritario H2 (verificare): ",
        "metric_total": "Aziende analizzate",
        "metric_eligible": "Aziende idonee (score > 0)",
        # Colonne template (la 2a è il codice di classificazione, qui ATECO)
        "cols_fase1": ["nome azienda", "codice ateco", "dimensione", "fatturato [M€]", "dipendenti", "ubicazione/consorzio", "vicinanza South H2 corridor", "AIA (si/no)", "consumo energia stimato [MWh]", "processo", "note"],
        "cols_fase2": ["nome azienda", "dimensione azienda", "codice ateco", "fabbisogno identificato [t/y]"],
        "esiti": {
            "spreco": "🔴 NON NECESSARIO: spreco termodinamico. Elettrificare, usare pompe di calore o biomasse.",
            "assoluto": "🟢 ASSOLUTAMENTE NECESSARIO: H2 richiesto chimicamente come materia prima / agente riducente.",
            "limiti": "🟢 NECESSARIO: difficile elettrificare per limiti fisici dei forni fusori.",
            "opzionale": "🟡 OPZIONALE: elevata competizione economica con Biometano e CSS.",
            "alert": "🟠 ALERT ELETTRIFICAZIONE: valutare forni elettrici/induzione prima dell'idrogeno.",
            "non_class": "⚪ NON CLASSIFICATO: processo a bassa temperatura o assente (elettrificabile).",
        },
    },
    "en": {
        "title": "🚀 H2READY TOOLKIT - Tool 2.1: HTA and RED III Company Scouting",
        "credits": "Developed within the [INTERREG H2Ready](https://www.ita-slo.eu/en/h2ready) project by **Matteo De Piccoli - [APE FVG](https://www.ape.fvg.it/)**",
        "instr_title": "📖 OPERATIONAL GUIDE (Read before starting)",
        "instructions_md": """
### 🎯 What is your goal?
As an Entity representative, your task is to **map the industries in your area** to understand which ones have a *real* need for green hydrogen (e.g. steelworks, glassworks, chemicals).

**Follow these 3 steps:**

**🟢 PHASE 1: Initial Screening**
Upload the screening file in the dedicated section. The simulator will filter eligible companies. If you don't have the file, download the template below the upload area.

**🟡 PHASE 2: Needs Consolidation**
Enter the estimated needs (tons/year) for the eligible companies. The template is available below the uploader.

**🔵 PHASE 3: Export**
Enter your **Identification Code** at the bottom of the page and click "Save".
        """,
        "note_codes": "**Note on NACE codes:** the tool reads only the **first 4 digits** of the code (the *Class*), ignoring the trailing digits used for purely statistical purposes. If you enter a code not in the hydrogen priority database, the tool still recognises the **macro-sector** (first 2 digits) and raises an alert asking you to verify the data.",
        "info_template1": "💡 If you don't have the screening data yet, use this template. **Warning: data must be uploaded keeping exactly this format (same column names and structure).**",
        "info_template2": "💡 Use this template to consolidate hydrogen tons. **Warning: data must be uploaded strictly following the format of this template.**",
        "btn_template1": "📥 Download Screening Template (Phase 1)",
        "btn_template2": "📥 Download Needs Template (Phase 2)",
        "header_fase1": "📤 PHASE 1: Upload and Thermodynamic Analysis",
        "header_fase2": "📤 PHASE 2: Needs Consolidation",
        "input_id": "Enter the Identification Code (e.g. 030043):",
        "btn_export": "🚀 Save Results to Central Database",
        "export_success": "✅ Data successfully saved!",
        "col_desc": "Sector description",
        "col_score": "H2 score",
        "col_verdict": "Thermodynamic verdict",
        "alert_codes": "⚠️ Codes outside the H2 priority database (please verify): ",
        "metric_total": "Companies analysed",
        "metric_eligible": "Eligible companies (score > 0)",
        "cols_fase1": ["company name", "nace code", "size", "turnover [M€]", "employees", "location/consortium", "proximity to South H2 corridor", "IED (yes/no)", "estimated energy consumption [MWh]", "process", "notes"],
        "cols_fase2": ["company name", "company size", "nace code", "identified need [t/y]"],
        "esiti": {
            "spreco": "🔴 NOT NECESSARY: thermodynamic waste. Electrify, use heat pumps or biomass.",
            "assoluto": "🟢 ABSOLUTELY NECESSARY: H2 required chemically as feedstock / reducing agent.",
            "limiti": "🟢 NECESSARY: difficult to electrify due to physical limits of melting furnaces.",
            "opzionale": "🟡 OPTIONAL: high economic competition with Biomethane and SRF.",
            "alert": "🟠 ELECTRIFICATION ALERT: evaluate electric/induction furnaces before hydrogen.",
            "non_class": "⚪ NOT CLASSIFIED: low-temperature or absent process (electrifiable).",
        },
    },
    "sl": {
        "title": "🚀 H2READY TOOLKIT - Orodje 2.1: Iskanje podjetij HTA in RED III",
        "credits": "Razvito v okviru projekta [INTERREG H2Ready](https://www.ita-slo.eu/en/h2ready), avtor **Matteo De Piccoli - [APE FVG](https://www.ape.fvg.it/)**",
        "instr_title": "📖 OPERATIVNI VODNIK (Preberite pred začetkom)",
        "instructions_md": """
### 🎯 Kakšen je vaš cilj?
Kot predstavnik organa je vaša naloga **kartiranje industrij na vašem območju**, da ugotovite, katere imajo *resnično* potrebo po zelenem vodiku (npr. jeklarne, steklarne, kemična industrija).

**Sledite tem 3 korakom:**

**🟢 1. FAZA: Začetni pregled**
Naložite datoteko za pregled v namenski razdelek. Simulator bo filtriral ustrezna podjetja. Če datoteke nimate, prenesite predlogo pod območjem za nalaganje.

**🟡 2. FAZA: Konsolidacija potreb**
Vnesite ocenjene potrebe (tone/leto) za ustrezna podjetja. Predloga je na voljo pod nalagalnikom.

**🔵 3. FAZA: Izvoz**
Na dnu strani vnesite svojo **Identifikacijsko kodo** in kliknite "Shrani".
        """,
        "note_codes": "**Opomba o kodah SKD:** orodje prebere le **prve 4 števke** kode (*razred*) in zanemari zadnje števke, ki imajo zgolj statistični namen. Če vnesete kodo, ki je ni v prednostni vodikovi bazi, orodje vseeno prepozna **makro-sektor** (prvi 2 števki) in sproži opozorilo, da preverite podatek.",
        "info_template1": "💡 Če še nimate podatkov za pregled, uporabite to predlogo. **Opozorilo: podatke je treba naložiti v natančno tem formatu (enaka imena stolpcev in struktura).**",
        "info_template2": "💡 Uporabite to predlogo za konsolidacijo ton vodika. **Opozorilo: podatke naložite strogo v skladu s formatom te predloge.**",
        "btn_template1": "📥 Prenesi predlogo za pregled (1. faza)",
        "btn_template2": "📥 Prenesi predlogo za potrebe (2. faza)",
        "header_fase1": "📤 1. FAZA: Nalaganje in termodinamična analiza",
        "header_fase2": "📤 2. FAZA: Konsolidacija potreb",
        "input_id": "Vnesite identifikacijsko kodo (npr. 030043):",
        "btn_export": "🚀 Shrani rezultate v centralno bazo",
        "export_success": "✅ Podatki so uspešno shranjeni!",
        "col_desc": "Opis sektorja",
        "col_score": "Ocena H2",
        "col_verdict": "Termodinamična ocena",
        "alert_codes": "⚠️ Kode izven prednostne baze H2 (preverite): ",
        "metric_total": "Analizirana podjetja",
        "metric_eligible": "Ustrezna podjetja (ocena > 0)",
        "cols_fase1": ["ime podjetja", "koda skd", "velikost", "promet [M€]", "zaposleni", "lokacija/konzorcij", "bližina South H2 corridor", "IED (da/ne)", "ocenjena poraba energije [MWh]", "proces", "opombe"],
        "cols_fase2": ["ime podjetja", "velikost podjetja", "koda skd", "identificirana potreba [t/y]"],
        "esiti": {
            "spreco": "🔴 NI POTREBNO: termodinamični odpadek. Elektrificirajte, uporabite toplotne črpalke ali biomaso.",
            "assoluto": "🟢 NUJNO POTREBNO: H2 je kemično potreben kot surovina / reducent.",
            "limiti": "🟢 POTREBNO: težko elektrificirati zaradi fizičnih omejitev talilnih peči.",
            "opzionale": "🟡 NEOBVEZNO: velika gospodarska konkurenca z biometanom in trdnimi alternativnimi gorivi (SRF).",
            "alert": "🟠 OPOZORILO O ELEKTRIFIKACIJI: ocenite električne/indukcijske peči pred vodikom.",
            "non_class": "⚪ NI RAZVRŠČENO: nizkotemperaturni ali odsoten proces (elektrificirljiv).",
        },
    },
}
_t = T[LANG]

# ==========================================
# 4. FUNZIONI DI CLASSIFICAZIONE (TRONCAMENTO A 4 CIFRE)
# ==========================================
def normalize_code(raw):
    """Toglie punti/virgole/spazi e tiene SOLO le prime 4 cifre.
    Vale identico per ATECO (24.10.00), NACE (24.10) e SKD (24.100)."""
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    return digits[:4]

def find_code_column(columns):
    """Trova la colonna del codice qualunque sia la lingua (ateco/nace/skd)."""
    for c in columns:
        cl = c.lower()
        if "ateco" in cl or "nace" in cl or "skd" in cl:
            return c
    return None

def describe_code(code4, lang):
    """Descrizione 'a cascata': 4 cifre -> macro 2 cifre -> alert.
    Ritorna (descrizione, is_in_db)."""
    desc_db = CODE_DESCRIPTIONS[lang]
    macro_db = MACRO_DESCRIPTIONS[lang]
    if code4 in desc_db:
        return desc_db[code4], True
    prefix = code4[:2]
    if prefix in macro_db:
        label = {"it": "Macro-settore", "en": "Macro-sector", "sl": "Makro-sektor"}[lang]
        return f"{label} {prefix}: {macro_db[prefix]}", False
    return {"it": "Codice non riconosciuto", "en": "Unrecognised code", "sl": "Neprepoznana koda"}[lang], False

def get_base_score(code4, testo_tecnico, lang):
    """Logica termodinamica + RED III. Ritorna (score_base, esito_text)."""
    esiti = T[lang]["esiti"]
    prefix = code4[:2]
    tt = testo_tecnico.lower()

    # 1. SPRECO TERMODINAMICO (Score 0)
    if prefix in ['35', '38', '41', '42', '43'] or code4.startswith('63') \
       or code4 == '2011' or code4 == '2013' or code4 == '1910':
        return 0, esiti["spreco"]

    # 2. ASSOLUTAMENTE NECESSARIO (Feedstock / agente riducente)
    if code4 in ['2015', '2014'] or code4 == '1920':
        if any(k in tt for k in ['etilen', 'ethylen', 'plastic', 'plastik']):
            return 0, esiti["spreco"]  # H2 di scarto da steam cracking
        return 5, esiti["assoluto"]
    if code4 == '2410' and any(k in tt for k in ['dri', 'riduzione diretta', 'direct reduction', 'neposredna redukcija']):
        return 5, esiti["assoluto"]

    # 3. NECESSARIO PER LIMITI FISICI (Vetro)
    if code4 in ['2311', '2313']:
        return 4.5, esiti["limiti"]

    # 4. OPZIONALE / COMPETIZIONE (Calcinazione)
    if code4 in ['2351', '2352', '2320', '2332']:
        return 3, esiti["opzionale"]

    # 5. ALERT ELETTRIFICAZIONE (Trattamenti termici / metallurgia)
    if code4 in ['2431', '2550', '2561', '2562'] or prefix == '24':
        return 2, esiti["alert"]

    # 6. Processo termico borderline (parole chiave)
    parole_chiave = ['metano', 'methane', 'metan', 'mw', 'forno', 'furnace', 'peč',
                     'fusione', 'melting', 'litje', 'calore', 'heat', 'toplota', 'ossidazione', 'oxidation']
    if any(k in tt for k in parole_chiave) and prefix in ['25', '26', '27', '28', '33']:
        return 1.5, esiti["alert"]

    return 0, esiti["non_class"]

def calculate_total_score(base, row, lang):
    """Applica i moltiplicatori dimensione + bonus (AIA/IED, zona industriale, corridoio)."""
    if base == 0:
        return 0
    cols = {c.lower(): c for c in row.index}

    def get(*keys):
        for k in keys:
            for cl, orig in cols.items():
                if k in cl:
                    return str(row.get(orig, "")).strip().lower()
        return ""

    dim = get('dimen', 'size', 'velikost')
    mult = 1.5 if dim in ['grande', 'large', 'velika'] else (1.2 if dim in ['media', 'medium', 'srednja'] else 1.0)
    score = base * mult

    yes_words = ['sì', 'si', 'yes', 'y', 'da']
    if get('aia', 'ied') in yes_words:
        score += 2
    ubic = get('ubic', 'locat', 'lokac')
    if "z.i." in ubic or ubic in yes_words:
        score += 3
    if get('south') in yes_words:
        score += 3
    return round(score, 1)

# ==========================================
# 5. INTESTAZIONE E CREDITI
# ==========================================
st.title(_t["title"])
st.markdown(_t["credits"])
st.markdown("""
    <p style='font-size: 0.8rem; color: gray;'>
        🌐 Progetto: <a href='https://www.ita-slo.eu/en/h2ready' target='_blank'>Interreg H2Ready</a> |
        🏠 Sito Ente: <a href='https://www.ape.fvg.it/' target='_blank'>APE FVG</a> |
        📧 Contatto: <a href='mailto:matteo.depiccoli@ape.fvg.it'>matteo.depiccoli@ape.fvg.it</a>
    </p>
""", unsafe_allow_html=True)
st.divider()

# ==========================================
# 6. ISTRUZIONI (MENU A TENDINA)
# ==========================================
with st.expander(_t["instr_title"], expanded=True):
    st.markdown(_t["instructions_md"])
    st.markdown("---")
    st.markdown(_t["note_codes"])
    st.markdown("---")
    nome_file_logica = f"logic_HTA_{LANG}.md"
    if os.path.exists(nome_file_logica):
        with open(nome_file_logica, "r", encoding="utf-8") as f:
            st.markdown(f.read())
    else:
        st.caption(f"ℹ️ File di logica ({nome_file_logica}) non trovato nella cartella.")

st.markdown("---")

# ==========================================
# 7. GENERATORI DI TEMPLATE (colonne dinamiche per lingua)
# ==========================================
# Righe d'esempio (language-neutral): coprono tutti i rami del motore di scoring.
# Formato: (nome, codice, taglia[L/M/S], fatturato, dipendenti, ubicazione, south, aia, energia, row_id)
EXAMPLE_ROWS = [
    ("Ferriere Isontine S.p.A.",       "24.10.00", "L", 320, 540, "Z.I. Aussa-Corno",   True,  True,  210000, "r1"),
    ("NitroFert FVG S.r.l.",           "20.15",    "L", 260, 410, "Z.I. Torviscosa",     True,  True,  175000, "r2"),
    ("Raffineria Adriatica S.p.A.",    "19.20",    "L", 900, 700, "Porto di Monfalcone", True,  True,  480000, "r3"),
    ("PetrolChimica Nord S.p.A.",      "20.14",    "L", 540, 620, "Z.I. Trieste",        False, True,  300000, "r4"),
    ("Vetreria Giuliana S.r.l.",       "23.11",    "L", 110, 230, "Z.I. Ronchi",         True,  False, 62000,  "r5"),
    ("Bottiglieria Carso S.p.A.",      "23.13",    "M", 70,  150, "Sgonico",             False, False, 38000,  "r6"),
    ("Cementi del Friuli S.p.A.",      "23.51",    "L", 180, 290, "Fanna",               False, True,  240000, "r7"),
    ("Calce e Dolomie Carnia S.r.l.",  "23.52",    "M", 45,  90,  "Tolmezzo",            False, False, 55000,  "r8"),
    ("Trafilerie Pordenonesi S.r.l.",  "24.31",    "M", 60,  130, "Z.I. Pordenone",      False, True,  28000,  "r9"),
    ("Metalli Speciali Tarvisio S.p.A.","24.44",   "M", 85,  160, "Z.I. Tarvisio",       True,  False, 41000,  "r10"),
    ("Acciaierie EAF Bassa S.p.A.",    "24.10.00", "L", 410, 600, "Z.I. Cervignano",     False, True,  260000, "r11"),
    ("TermoVapore Servizi S.r.l.",     "35.30",    "M", 40,  75,  "Udine",               False, False, 90000,  "r12"),
    ("DataNord Cloud S.p.A.",          "63.11",    "L", 150, 95,  "Z.I. Amaro",          False, False, 120000, "r13"),
    ("GasTecnici Industriali S.p.A.",  "20.11",    "L", 130, 180, "Z.I. Aussa-Corno",    True,  True,  140000, "r14"),
    ("MicroElettronica Maniago S.r.l.","26.01",    "S", 12,  35,  "Z.I. Maniago",        False, False, 4200,   "r15"),
]

# Valori categorici tradotti (devono combaciare con la logica di scoring)
SIZE_LABEL = {"it": {"L": "Grande", "M": "Media", "S": "Piccola"},
              "en": {"L": "Large", "M": "Medium", "S": "Small"},
              "sl": {"L": "Velika", "M": "Srednja", "S": "Mala"}}
YESNO = {"it": {True: "Sì", False: "No"},
         "en": {True: "Yes", False: "No"},
         "sl": {True: "Da", False: "Ne"}}

# Testo colonna "processo" per riga e lingua (le parole chiave sono multilingua)
PROC_TEXT = {
    "r1":  {"it": "Acciaio primario - Ciclo DRI a idrogeno", "en": "Primary steel - hydrogen DRI cycle", "sl": "Primarno jeklo - vodikov cikel DRI"},
    "r2":  {"it": "Sintesi ammoniaca / fertilizzanti azotati", "en": "Ammonia synthesis / nitrogen fertilizers", "sl": "Sinteza amonijaka / dušikova gnojila"},
    "r3":  {"it": "Idrodesolforazione e hydrocracking", "en": "Hydrodesulfurization and hydrocracking", "sl": "Hidrorazžveplanje in hidrokreking"},
    "r4":  {"it": "Steam cracking per etilene e plastica", "en": "Steam cracking for ethylene and plastics", "sl": "Parni kreking za etilen in plastiko"},
    "r5":  {"it": "Forni fusori continui vetro piano >100 t/g", "en": "Continuous flat glass melting furnaces >100 t/d", "sl": "Kontinuirne talilne peči za ravno steklo >100 t/dan"},
    "r6":  {"it": "Forno fusorio vetro cavo", "en": "Hollow glass melting furnace", "sl": "Talilna peč za votlo steklo"},
    "r7":  {"it": "Forno rotativo clinker - calcinazione", "en": "Rotary kiln clinker - calcination", "sl": "Rotacijska peč za klinker - kalcinacija"},
    "r8":  {"it": "Calcinazione calcare in forni verticali", "en": "Limestone calcination in vertical kilns", "sl": "Kalcinacija apnenca v vertikalnih pečeh"},
    "r9":  {"it": "Trafilatura a freddo e ricottura", "en": "Cold drawing and annealing", "sl": "Hladno vlečenje in žarjenje"},
    "r10": {"it": "Fusione metalli non ferrosi", "en": "Non-ferrous metal melting", "sl": "Taljenje neželeznih kovin"},
    "r11": {"it": "Forno elettrico ad arco (EAF) da rottame", "en": "Electric arc furnace (EAF) from scrap", "sl": "Elektroobločna peč (EAF) iz odpadkov"},
    "r12": {"it": "Produzione e distribuzione vapore industriale", "en": "Industrial steam production and distribution", "sl": "Proizvodnja in distribucija industrijske pare"},
    "r13": {"it": "Data center - raffreddamento e UPS", "en": "Data center - cooling and UPS", "sl": "Podatkovni center - hlajenje in UPS"},
    "r14": {"it": "Produzione gas industriali via SMR", "en": "Industrial gas production via SMR", "sl": "Proizvodnja industrijskih plinov prek SMR"},
    "r15": {"it": "Saldatura e forno di reflow", "en": "Welding and reflow furnace", "sl": "Varjenje in reflow peč"},
}

# Testo colonna "note": qui indica l'esito atteso di ogni riga d'esempio
NOTE_TEXT = {
    "r1":  {"it": "Assoluto: H2 agente riducente", "en": "Absolute: H2 reducing agent", "sl": "Nujno: H2 kot reducent"},
    "r2":  {"it": "Assoluto: H2 feedstock RED III", "en": "Absolute: H2 feedstock RED III", "sl": "Nujno: H2 surovina RED III"},
    "r3":  {"it": "Assoluto: H2 di processo", "en": "Absolute: process H2", "sl": "Nujno: procesni H2"},
    "r4":  {"it": "Spreco: H2 di scarto (esclusione)", "en": "Waste: by-product H2 (exclusion)", "sl": "Odpadek: stranski H2 (izključitev)"},
    "r5":  {"it": "Limiti fisici: difficile elettrificare", "en": "Physical limits: hard to electrify", "sl": "Fizične omejitve: težko elektrificirati"},
    "r6":  {"it": "Limiti fisici", "en": "Physical limits", "sl": "Fizične omejitve"},
    "r7":  {"it": "Opzionale: compete con CSS/biometano", "en": "Optional: competes with SRF/biomethane", "sl": "Neobvezno: konkurenca SRF/biometan"},
    "r8":  {"it": "Opzionale", "en": "Optional", "sl": "Neobvezno"},
    "r9":  {"it": "Alert: valutare induzione elettrica", "en": "Alert: consider electric induction", "sl": "Opozorilo: razmislite o elektriki/indukciji"},
    "r10": {"it": "Alert: metallurgia generica (prefix 24)", "en": "Alert: generic metallurgy (prefix 24)", "sl": "Opozorilo: splošna metalurgija (predpona 24)"},
    "r11": {"it": "Alert: forno EAF, metallurgia prefix 24", "en": "Alert: EAF furnace, metallurgy prefix 24", "sl": "Opozorilo: peč EAF, metalurgija predpona 24"},
    "r12": {"it": "Spreco: elettrificare/biomasse", "en": "Waste: electrify/biomass", "sl": "Odpadek: elektrifikacija/biomasa"},
    "r13": {"it": "Spreco: elettricità pura", "en": "Waste: pure electricity", "sl": "Odpadek: čista elektrika"},
    "r14": {"it": "Spreco come input: SMR da sostituire", "en": "Waste as input: SMR to be replaced", "sl": "Odpadek kot vhod: SMR za zamenjavo"},
    "r15": {"it": "Codice fuori DB + parola termica -> verifica", "en": "Code outside DB + thermal keyword -> verify", "sl": "Koda izven baze + toplotna beseda -> preveri"},
}

# Testi del foglio Legenda
LEGEND = {
    "it": {
        "head": ["Colonna", "Valori ammessi / Note"],
        "rows": [
            ("nome azienda", "Testo libero (obbligatorio)"),
            ("codice", "Codice ATECO/NACE/SKD. Il tool legge solo le prime 4 cifre (es. 24.10.00 -> 2410). Obbligatorio."),
            ("dimensione", "Piccola / Media / Grande  (moltiplicatore ×1,0 / ×1,2 / ×1,5). Obbligatorio."),
            ("fatturato [M€]", "Numero (facoltativo)"),
            ("dipendenti", "Numero (facoltativo)"),
            ("ubicazione/consorzio", "Se contiene 'Z.I.' o = Sì -> bonus +3"),
            ("vicinanza South H2 corridor", "Sì / No  (Sì -> bonus +3)"),
            ("AIA (si/no)", "Sì / No  (Sì -> bonus +2)"),
            ("consumo energia [MWh]", "Numero (facoltativo)"),
            ("processo", "Testo libero: parole come 'DRI', 'forno', 'fusione' aiutano la classificazione"),
            ("note", "Testo libero. In questo template indica l'esito atteso di ogni riga d'esempio"),
        ],
        "head2": ["Esito", "Significato"],
        "rows2": [
            ("🟢 Assolutamente necessario", "H2 come materia prima / agente riducente (2015, 2014, 1920, 2410+DRI)"),
            ("🟢 Necessario (limiti fisici)", "Grandi forni fusori vetro (2311, 2313)"),
            ("🟡 Opzionale", "Cemento/calce/refrattari/laterizi (2351, 2352, 2320, 2332)"),
            ("🟠 Alert elettrificazione", "Trattamenti termici e metallurgia (2431, 2550, 2561, 2562, prefix 24)"),
            ("🔴 Spreco termodinamico", "Vapore/energia, edilizia, data center, H2 di scarto (35, 38, 41-43, 63, 2011, 2013, 1910)"),
        ],
    },
    "en": {
        "head": ["Column", "Allowed values / Notes"],
        "rows": [
            ("company name", "Free text (mandatory)"),
            ("code", "ATECO/NACE/SKD code. The tool reads only the first 4 digits (e.g. 24.10 -> 2410). Mandatory."),
            ("size", "Small / Medium / Large  (multiplier ×1.0 / ×1.2 / ×1.5). Mandatory."),
            ("turnover [M€]", "Number (optional)"),
            ("employees", "Number (optional)"),
            ("location/consortium", "If it contains 'Z.I.' or = Yes -> bonus +3"),
            ("proximity to South H2 corridor", "Yes / No  (Yes -> bonus +3)"),
            ("IED (yes/no)", "Yes / No  (Yes -> bonus +2)"),
            ("energy consumption [MWh]", "Number (optional)"),
            ("process", "Free text: words like 'DRI', 'furnace', 'melting' help classification"),
            ("notes", "Free text. In this template it shows the expected verdict of each example row"),
        ],
        "head2": ["Verdict", "Meaning"],
        "rows2": [
            ("🟢 Absolutely necessary", "H2 as feedstock / reducing agent (2015, 2014, 1920, 2410+DRI)"),
            ("🟢 Necessary (physical limits)", "Large glass melting furnaces (2311, 2313)"),
            ("🟡 Optional", "Cement/lime/refractories/bricks (2351, 2352, 2320, 2332)"),
            ("🟠 Electrification alert", "Heat treatments and metallurgy (2431, 2550, 2561, 2562, prefix 24)"),
            ("🔴 Thermodynamic waste", "Steam/energy, construction, data centers, by-product H2 (35, 38, 41-43, 63, 2011, 2013, 1910)"),
        ],
    },
    "sl": {
        "head": ["Stolpec", "Dovoljene vrednosti / Opombe"],
        "rows": [
            ("ime podjetja", "Prosto besedilo (obvezno)"),
            ("koda", "Koda ATECO/NACE/SKD. Orodje prebere le prve 4 števke (npr. 24.100 -> 2410). Obvezno."),
            ("velikost", "Mala / Srednja / Velika  (množitelj ×1,0 / ×1,2 / ×1,5). Obvezno."),
            ("promet [M€]", "Število (neobvezno)"),
            ("zaposleni", "Število (neobvezno)"),
            ("lokacija/konzorcij", "Če vsebuje 'Z.I.' ali = Da -> bonus +3"),
            ("bližina South H2 corridor", "Da / Ne  (Da -> bonus +3)"),
            ("IED (da/ne)", "Da / Ne  (Da -> bonus +2)"),
            ("poraba energije [MWh]", "Število (neobvezno)"),
            ("proces", "Prosto besedilo: besede kot 'DRI', 'peč', 'litje' pomagajo pri razvrstitvi"),
            ("opombe", "Prosto besedilo. V tej predlogi prikazuje pričakovano oceno vsake vrstice"),
        ],
        "head2": ["Ocena", "Pomen"],
        "rows2": [
            ("🟢 Nujno potrebno", "H2 kot surovina / reducent (2015, 2014, 1920, 2410+DRI)"),
            ("🟢 Potrebno (fizične omejitve)", "Velike talilne peči za steklo (2311, 2313)"),
            ("🟡 Neobvezno", "Cement/apno/ognjevzdržni izdelki/opeka (2351, 2352, 2320, 2332)"),
            ("🟠 Opozorilo o elektrifikaciji", "Toplotne obdelave in metalurgija (2431, 2550, 2561, 2562, predpona 24)"),
            ("🔴 Termodinamični odpadek", "Para/energija, gradbeništvo, podatkovni centri, odpadni H2 (35, 38, 41-43, 63, 2011, 2013, 1910)"),
        ],
    },
}


def _build_workbook(cols, data_rows, lang, dropdown_map):
    """Costruisce un workbook formattato con header, bordi, dropdown e foglio Legenda."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    note_font = Font(name="Arial", size=10, italic=True, color="7F7F7F")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border
    ws.row_dimensions[1].height = 32

    text_left_cols = {1, 6, 10}   # nome, ubicazione, processo
    note_col = len(cols)          # ultima colonna = note
    for r_idx, row in enumerate(data_rows, start=2):
        ws.append(row)
        for c_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            if c_idx == note_col:
                cell.alignment, cell.font = left, note_font
            elif c_idx in text_left_cols:
                cell.alignment, cell.font = left, body_font
            else:
                cell.alignment, cell.font = center, body_font

    widths = [26, 13, 12, 13, 11, 20, 18, 12, 16, 32, 32]
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i - 1 < len(widths) else 16
    for r in range(2, len(data_rows) + 2):
        for col in (4, 5, 9):
            ws.cell(row=r, column=col).number_format = '#,##0'
    ws.freeze_panes = "A2"

    # Menu a tendina
    last = len(data_rows) + 1
    for col_idx, values in dropdown_map.items():
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(col_idx)
        dv.add(f"{letter}2:{letter}{last}")

    # Foglio Legenda
    leg = LEGEND[lang]
    ws2 = wb.create_sheet("Legenda")
    ws2.append(leg["head"])
    for name, desc in leg["rows"]:
        ws2.append([name, desc])
    ws2.append(["", ""])
    sep_row = ws2.max_row + 1
    ws2.append(leg["head2"])
    for name, desc in leg["rows2"]:
        ws2.append([name, desc])
    for c in (1, 2):
        for hr in (1, sep_row):
            ws2.cell(row=hr, column=c).fill = header_fill
            ws2.cell(row=hr, column=c).font = header_font
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 85
    for r in range(1, ws2.max_row + 1):
        ws2.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        if ws2.cell(row=r, column=2).font.color is None or ws2.cell(row=r, column=2).font.b is not True:
            ws2.cell(row=r, column=2).font = Font(name="Arial", size=10)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_template_fase1(lang):
    cols = T[lang]["cols_fase1"]
    data_rows = []
    for name, code, size, turn, emp, loc, south, aia, energy, rid in EXAMPLE_ROWS:
        data_rows.append([
            name, code, SIZE_LABEL[lang][size], turn, emp, loc,
            YESNO[lang][south], YESNO[lang][aia], energy,
            PROC_TEXT[rid][lang], NOTE_TEXT[rid][lang],
        ])
    # Dropdown: col 3 dimensione, col 7 South corridor, col 8 AIA/IED
    dd = {
        3: list(SIZE_LABEL[lang].values()),
        7: [YESNO[lang][True], YESNO[lang][False]],
        8: [YESNO[lang][True], YESNO[lang][False]],
    }
    return _build_workbook(cols, data_rows, lang, dd)


def generate_template_fase2(lang):
    cols = T[lang]["cols_fase2"]
    # 3 esempi idonei coerenti con la Fase 1
    f2 = [
        ("Ferriere Isontine S.p.A.", "L", "24.10.00", 1800.0),
        ("NitroFert FVG S.r.l.", "L", "20.15", 2500.0),
        ("Vetreria Giuliana S.r.l.", "L", "23.11", 600.0),
    ]
    output = BytesIO()
    rows = [[n, SIZE_LABEL[lang][s], c, v] for n, s, c, v in f2]
    df_temp = pd.DataFrame(rows, columns=cols)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_temp.to_excel(writer, index=False)
    return output.getvalue()

# ==========================================
# 8. FASE 1: SCREENING + ANALISI
# ==========================================
st.header(_t["header_fase1"])
uploaded_file_1 = st.file_uploader("Upload Fase 1", type=["xlsx", "csv"], key="fase1")

if uploaded_file_1:
    try:
        df1 = pd.read_excel(uploaded_file_1) if uploaded_file_1.name.endswith('.xlsx') else pd.read_csv(uploaded_file_1)
        code_col = find_code_column(df1.columns)
        if code_col is None:
            st.error(f"❌ Colonna codice ({CLASSIF_NAME[LANG]}) non trovata nel file.")
        else:
            # Individua colonne testuali per le parole chiave di processo
            proc_cols = [c for c in df1.columns if any(k in c.lower() for k in ['process', 'proces', 'not', 'opomb'])]

            descrizioni, scores, esiti_out, codici_fuori_db = [], [], [], []
            for _, row in df1.iterrows():
                code4 = normalize_code(row.get(code_col, ""))
                testo = " ".join(str(row.get(c, "")) for c in proc_cols)
                desc, in_db = describe_code(code4, LANG)
                base, esito = get_base_score(code4, testo, LANG)
                total = calculate_total_score(base, row, LANG)
                descrizioni.append(desc)
                esiti_out.append(esito)
                scores.append(total)
                if not in_db and code4:
                    codici_fuori_db.append(code4)

            df1[_t["col_desc"]] = descrizioni
            df1[_t["col_score"]] = scores
            df1[_t["col_verdict"]] = esiti_out

            st.success("✅ File Fase 1 analizzato!")
            c1, c2 = st.columns(2)
            c1.metric(_t["metric_total"], len(df1))
            c2.metric(_t["metric_eligible"], int((df1[_t["col_score"]] > 0).sum()))

            if codici_fuori_db:
                st.warning(_t["alert_codes"] + ", ".join(sorted(set(codici_fuori_db))))

            st.dataframe(df1.sort_values(_t["col_score"], ascending=False), use_container_width=True)
    except Exception as e:
        st.error(f"Errore: {e}")

st.info(_t["info_template1"])
st.download_button(_t["btn_template1"], generate_template_fase1(LANG), "template_screening.xlsx")

st.markdown("---")

# ==========================================
# 9. FASE 2: FABBISOGNI + DOWNLOAD TEMPLATE 2
# ==========================================
st.header(_t["header_fase2"])
uploaded_file_2 = st.file_uploader("Upload Fase 2", type=["xlsx", "csv"], key="fase2")

n_aziende = 0
totale_h2 = 0.0
nomi_aziende = ""

if uploaded_file_2:
    try:
        df2 = pd.read_excel(uploaded_file_2) if uploaded_file_2.name.endswith('.xlsx') else pd.read_csv(uploaded_file_2)
        df2.columns = df2.columns.str.strip().str.lower()
        st.success("✅ Dati Fabbisogni caricati!")
        st.dataframe(df2, use_container_width=True)

        # Colonna fabbisogno e colonna nome rilevate in modo language-agnostic
        col_target = next((c for c in df2.columns if 'fabbisogno' in c or 'need' in c or 'potreb' in c), None)
        col_nome = next((c for c in df2.columns if 'azienda' in c or 'company' in c or 'podjet' in c), None)
        if col_target:
            totale_h2 = pd.to_numeric(df2[col_target], errors='coerce').fillna(0).sum()
            n_aziende = int((pd.to_numeric(df2[col_target], errors='coerce').fillna(0) > 0).sum())
            if col_nome:
                nomi_aziende = "; ".join(df2[col_nome].astype(str).tolist())
            st.metric("H2 ton/anno", f"{totale_h2:,.1f}")
    except Exception as e:
        st.error(f"Errore: {e}")

st.info(_t["info_template2"])
st.download_button(_t["btn_template2"], generate_template_fase2(LANG), "template_fabbisogni.xlsx")

# ==========================================
# 10. ESPORTAZIONE (Codice Identificativo)
# ==========================================
st.divider()
st.subheader("🔗 Esportazione Dati")
id_identificativo = st.text_input(_t["input_id"])

if st.button(_t["btn_export"]):
    if not id_identificativo:
        st.error("Inserisci il Codice Identificativo!")
    else:
        payload = {
            "ID_ISTAT": id_identificativo,
            "T21_N_AZIENDE_IDONEE": n_aziende,
            "T21_NOMI_AZIENDE": nomi_aziende,
            "T21_FABBISOGNO_H2_TON_ANNO": round(float(totale_h2), 2)
        }
        GOOGLE_URL = "https://script.google.com/macros/s/AKfycbwpP0x0hBnhOadXA43IieWg9EusAuhaafpyeXpyaStssDd7Qo-jwnuOttAllzz8r5JS/exec"
        try:
            response = requests.post(GOOGLE_URL, data=json.dumps(payload), headers={'Content-Type': 'application/json'})
            if response.status_code == 200:
                st.success(_t["export_success"])
                st.balloons()
        except Exception:
            st.error("Errore di connessione.")
